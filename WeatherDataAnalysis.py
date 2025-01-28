
"""
Weather Data Analysis Program 
Essentially creating a program to analyze weather from an Excel file as long as it's located in the
domain of python. Because python would cross - check the users input to see if it has access to both take
data from it and modify the Excel sheet.
1) I'll need 2 bar charts (i) for avg monthly temp and 
                            (ii) for percipitation analysis
2) A vizualization for both - if with enough time probably let the users have a say ...
            on the colour and more lackluster things like that
"""


import openpyxl
from openpyxl.chart import Reference, BarChart
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook


''' Introducing a def function to load file_name which would be user input-ed'''
# Due to the nature of a def function and everything indented underneath. This block of code
        # is associated with what the user as input-ed to see if it's within the workbook if it is
            # It's then activated and would return the file if not (using an f string) we pass the information of it
                # not being the directory.
def load_excel(file_name):

    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        return workbook, sheet
    except FileNotFoundError:
        print(f"Error: File '{file_name}' not found in the directory.")
        exit()


''' This 2nd Def function is to analyze monthly Avg temp and determine the month with the highest avg'''
''' Thanks to Youtube and free code Camp'''
def analyze_temperature(sheet):

    temp_data = {} # An empty dictionary for which is associated with the variable temp_data and can be appended
                    # if need be
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skips header row. Extra - Extra help from You-tube.
        date, temp = row[0], row[1]
        if date and temp is not None:
            month = date.strftime("B")
            temp_data.setdefault(month, []).append(temp)  # A function .setdefault associated with the empty dict
                                                        # and is further appended with temp which has being defined Earlier

    # calculating monthly averages using basic aretmatic and dictionary.
    avg_temps = {month: sum(values) / len(values) for month, values in temp_data.items()}   # values in the bracket of the function is adjecent with Excel on python
    highest_month = max(avg_temps, key=avg_temps.get) # using the get function that's attached to a variable that has stored data and making sure it's the highest with (max
    print(
        f"The month with the highest average temperature is {highest_month} with an average of {avg_temps[highest_month]:.2f}°C.")

    return avg_temps, highest_month
        # Returning 2 values that are attached to 2 variables that have been defined above (sequential execution)


        # Save the average temp and make the bar chart
def save_temperature_analysis(workbook, avg_temps): # placing workbook and avg_temp variables in this def fuction are the 2
                                                    # x and y components we need to make the bar chart
    sheet = workbook.create_sheet("AverageTemperatureAnalysis")
    sheet.append(["Month", "Average Temperature (°C)"])

    for month, avg in avg_temps.items():
        sheet.append([month, avg])

    # Bar Chart customization
    chart = BarChart()
    chart.title = "Average Monthly Temperatures"
    chart.x_axis.title = "Month"
    chart.y_axis.title = "Temperature (°C)"

    '''More youtube help'''
    #TerribleWithGridMaking
    data = Reference(sheet, min_col=2, min_row=2, max_row=len(avg_temps) + 1)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=len(avg_temps) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    sheet.add_chart(chart, "D2")


'''Another def function for analyzing percipitation while passing sheet as an argument 
remembering sheet contains activating the work book'''
def analyze_precipitation(sheet):

    precip_data = {} #Creating an empty dictionary to store/append values to and get.
    ''' More youtube help '''
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        date, precip = row[0], row[2]
        if date and precip is not None:
            precip_data[date] = precip

    highest_day = max(precip_data, key=precip_data.get)
    print(f"The day with the highest precipitation is {highest_day} with {precip_data[highest_day]:.2f} mm.")

    return precip_data, highest_day


 # Save the average temp and make the bar chart - Essentially doing what we did for the temperature.
def save_precipitation_analysis(workbook, precip_data):

    sheet = workbook.create_sheet("PrecipitationAnalysis")
    sheet.append(["Date", "Precipitation (mm)"])

    for date, precip in precip_data.items():
        sheet.append([date, precip])

    # Bar Chart customization for precipitation this time Same proccess.
    chart = BarChart()
    chart.title = "Daily Precipitation"
    chart.x_axis.title = "Date"
    chart.y_axis.title = "Precipitation (mm)"

    data = Reference(sheet, min_col=2, min_row=2, max_row=len(precip_data) + 1)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=len(precip_data) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    sheet.add_chart(chart, "D2")


''' Main function stating all our variables and everything that would be called on later
    When defining functions and creating bar-charts.'''
def main():

    # User input for file name
    file_name = input("Enter the name of the Excel file/sheet \nExample: WeatherDataAnalysis.xlsx.xlsx \n:  ")

    # Load the data
    workbook, sheet = load_excel(file_name)

    # Analyze temperature
    avg_temps, highest_month = analyze_temperature(sheet)
    save_temperature_analysis(workbook, avg_temps)

    # Analyze precipitation
    precip_data, highest_day = analyze_precipitation(sheet)
    save_precipitation_analysis(workbook, precip_data)

    # Save results to a new file
    output_file = "WeatherDataAnalysis.xlsx"
    workbook.save(output_file)
    print(f"Analysis complete. Results saved to '{output_file}'.")


main()
